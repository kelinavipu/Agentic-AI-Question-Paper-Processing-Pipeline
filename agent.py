import os
import re
import cv2
import json
import numpy as np
import pytesseract
import operator
from typing import TypedDict, Annotated, List, Dict, Any, Optional
from pathlib import Path

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

import time
import threading
import concurrent.futures
from fpdf import FPDF

from langchain_core.messages import HumanMessage, AIMessage
from langchain_core.tools import tool
from langgraph.graph import StateGraph, END
from langgraph.checkpoint.memory import MemorySaver
from langchain_groq import ChatGroq

# Load .env from project root
try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).parent / ".env"
    load_dotenv(dotenv_path=_env_path)
    print(f"[Extracta Agent] Loaded .env from {_env_path}")
except ImportError:
    pass  # python-dotenv not installed, fall back to system env

# --------------------------------------------------------------------------
# CONFIG & TESSERACT PATH RESOLUTION
# --------------------------------------------------------------------------

def _find_tesseract() -> str:
    candidates = [
        r"C:\Users\Admin\AppData\Local\Programs\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files\tesseract.exe",
        r"C:\Users\Kelina\AppData\Local\Programs\Tesseract-OCR\tesseract.exe",
    ]
    for p in candidates:
        if os.path.exists(p):
            print(f"[Extracta Agent] Found Tesseract at: {p}")
            return p
    print("[Extracta Agent] Tesseract not found in standard paths, using PATH fallback.")
    return "tesseract"

pytesseract.pytesseract.tesseract_cmd = _find_tesseract()

def build_llm(api_key: str, temp: float = 0.0, model_name: str = "llama-3.3-70b-versatile"):
    if not api_key:
        return None
    return ChatGroq(model=model_name, temperature=temp, api_key=api_key)

# Summary / Spark LLM key (GROQ_SUMMARY_KEY in .env)
GROQ_SUMMARY_KEY = (
    os.environ.get("GROQ_SUMMARY_KEY")
    or os.environ.get("GROQ_API_KEY", "")
)

if GROQ_SUMMARY_KEY:
    LLM = build_llm(GROQ_SUMMARY_KEY)
    LLM_AVAILABLE = True
    print("[Extracta Agent] LLM initialized successfully.")
else:
    LLM = None
    LLM_AVAILABLE = False
    print("[Extracta Agent] WARNING: LLM initialization failed: No API Key found.")

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

OCR_CONFIG  = r"--oem 3 --psm 6"
MAX_RETRIES = 3

task_progress_logs: Dict[str, List[Dict]] = {}

def update_task_progress(task_id: str, node_name: str, percentage: int, status_message: str):
    if task_id not in task_progress_logs:
        task_progress_logs[task_id] = []
    task_progress_logs[task_id].append({
        "node": node_name,
        "progress": percentage,
        "message": status_message
    })
    print(f"[{task_id}] Node: {node_name} | Progress: {percentage}% | {status_message}")


# --------------------------------------------------------------------------
# SHARED STATE SCHEMA
# --------------------------------------------------------------------------

class QPState(TypedDict):
    task_id:        str
    university:     str                # e.g. "mumbai" or "generic"
    pdf_path:       str
    image_paths:    List[str]
    raw_text:       str
    clean_text:     str
    header:         Dict[str, str]
    body_text:      str
    normalized:     str
    structured:     Dict[str, Any]
    diagram_paths:  List[str]
    flat_rows:      List[List[str]]
    excel_path:     str
    validation_ok:  bool
    retry_count:    int
    messages:       Annotated[List[Any], operator.add]
    errors:         List[str]


# --------------------------------------------------------------------------
# HEADER EXTRACTION — UNIVERSITY-AWARE
# --------------------------------------------------------------------------

def _extract_header_generic(text: str) -> Dict[str, str]:
    h = dict(paper_code="", subject="", date="", exam="",
             max_marks="", time="", qp_code="",
             semester="", scheme="", branch="", note="")
    if m := re.search(r"Subject\s*Code[:\s]+([A-Za-z0-9\-]+)", text, re.I):
        h["paper_code"] = m.group(1).strip()
    if m := re.search(r"Subject\s*Code[:\s]+[A-Za-z0-9\-]+\s*/\s*(.*?)(?:\n|\d{2}/\d{2}/\d{4})", text, re.I):
        h["subject"] = m.group(1).strip().rstrip("/").strip()
    if m := re.search(r"\d{2}/\d{2}/\d{4}", text):
        h["date"] = m.group()
    if m := re.search(r"Max\.?\s*Marks[:\s]+(\d+)", text, re.I):
        h["max_marks"] = m.group(1)
    if m := re.search(r"Time[:\s]+([0-9\.]+\s*hours?)", text, re.I):
        h["time"] = m.group(1)
    if m := re.search(r"QP\s*CODE[:\s]+(\d+)", text, re.I):
        h["qp_code"] = m.group(1)
    if m := re.search(r"Note[:\s]+(.+?)(?:\n|$)", text, re.I):
        h["note"] = m.group(1).strip()
    return h


def _extract_header_mumbai(text: str) -> Dict[str, str]:
    """
    Tuned for Mumbai University question paper format:
    Paper / Subject Code: 37473 / Software Engineering and Project Management
    10/12/2024 CSE-AIML SEM-VI C SCHEME SEPM QP CODE: 10064710
    Time: 3 hour   Max. Marks: 80
    Note: Question 1 is compulsory. Attempt any 3 out of remaining 5 questions.
    """
    h = dict(paper_code="", subject="", date="", exam="",
             max_marks="", time="", qp_code="",
             semester="", scheme="", branch="", note="")

    # Paper code & subject name from "Paper / Subject Code: XXXXX / Subject Name"
    psc = re.search(
        r"(?:Paper\s*/\s*)?Subject\s*Code[:\s]+([A-Za-z0-9\-]+)\s*/\s*(.*?)(?:\n|$)",
        text, re.I
    )
    if psc:
        h["paper_code"] = psc.group(1).strip()
        h["subject"]    = psc.group(2).strip().rstrip("/").strip()

    # Date — DD/MM/YYYY
    if m := re.search(r"\d{2}/\d{2}/\d{4}", text):
        h["date"] = m.group()

    # Branch e.g. CSE-AIML, CE, IT, EXTC
    if m := re.search(r"\b(CSE[-\s]?[A-Z]+|CE|IT|EXTC|MECH|CIVIL|AIDS|DS)\b", text, re.I):
        h["branch"] = m.group(1).upper()

    # Semester e.g. SEM-VI, SEM V, SEMVI
    if m := re.search(r"SEM[-\s]?([IVX]+|\d+)", text, re.I):
        h["semester"] = "SEM-" + m.group(1).strip().upper()

    # Scheme e.g. C SCHEME, C-SCHEME, D SCHEME
    if m := re.search(r"([A-D])[-\s]?SCHEME", text, re.I):
        h["scheme"] = m.group(1).upper() + " SCHEME"

    # QP Code
    if m := re.search(r"QP\s*CODE[:\s]+(\d+)", text, re.I):
        h["qp_code"] = m.group(1)

    # Time
    if m := re.search(r"Time[:\s]+([0-9\.]+\s*(?:hours?|hrs?))", text, re.I):
        h["time"] = m.group(1).strip()

    # Max Marks
    if m := re.search(r"Max\.?\s*Marks[:\s]+(\d+)", text, re.I):
        h["max_marks"] = m.group(1)

    # Note / Instructions
    if m := re.search(r"Note[:\s]+(.+?)(?:\n(?:Q|$)|$)", text, re.I | re.DOTALL):
        raw_note = m.group(1).strip().replace("\n", " ")
        h["note"] = raw_note[:300]

    return h


def _extract_header_abvv(text: str) -> Dict[str, str]:
    """
    Tuned for ABVV (Atal Bihari Vajpayee Vishwavidyalaya) exam paper format.
    """
    h = dict(
        paper_code="", subject="", date="", exam="",
        max_marks="", time="", qp_code="",
        semester="", scheme="", branch="",
        exam_session="", paper_type="", note=""
    )

    # SI code / Paper code — e.g. SI-003457
    if m := re.search(r"\b(SI[-\s]?\d{4,8})\b", text, re.I):
        h["paper_code"] = m.group(1).strip().upper()
        h["qp_code"] = h["paper_code"]

    # Branch
    if m := re.search(r"(?:B\.A\.|B\.Sc\.|B\.Com|B\.E\.|B\.Tech|M\.A\.|M\.Sc\.|M\.Com)[^\n]+", text, re.I):
        h["branch"] = m.group(0).strip().rstrip('/')

    # Subject — e.g. (G01) or (GO1) ELEMENTARY BOTANY (BOTANY)
    if m := re.search(r"\([A-Z][A-Z0-9]*\)\s+([A-Za-z &\'\-\(\)]+)", text):
        h["subject"] = m.group(1).strip()
    elif m := re.search(r"Concept of Business|Business\s+\w+|Management\s+\w+", text, re.I):
        h["subject"] = m.group(0).strip()

    # Semester — First Semester, End Semester, etc.
    if m := re.search(r"(First|Second|Third|Fourth|Fifth|Sixth|End|\d+(?:st|nd|rd|th))\s+Semester", text, re.I):
        h["semester"] = m.group(0).strip()

    # Exam session — e.g. December-2025-26
    if m := re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)[-\s]\d{4}(?:[-–]\d{2,4})?", text, re.I):
        h["exam_session"] = m.group(0).strip()
        h["date"] = h["exam_session"]

    # Exam type — Regular/PVT, etc.
    if m := re.search(r"\((Regular(?:\s*/\s*PVT)?|Ex[-\s]?Regular|Private|PVT)\)", text, re.I):
        h["exam"] = m.group(1).strip()

    # Paper type / Scheme
    if m := re.search(r"(Compulsory\s*/\s*Optional|Compulsory|Optional)", text, re.I):
        h["paper_type"] = m.group(1).strip()
        h["scheme"] = h["paper_type"]
    if m := re.search(r"Course type\s*[:\-]?\s*([A-Za-z0-9]+)", text, re.I):
        h["scheme"] = m.group(1).strip()

    # Time
    if m := re.search(r"Time\s*[:\-]?\s*([A-Za-z\d\.]+(?:\s+Hours?|\s+Hrs?))", text, re.I):
        h["time"] = m.group(1).strip()

    # Max Marks
    if m := re.search(r"Max(?:imum)?[.\s]*Marks\s*[:\-]?\s*(\d+)", text, re.I):
        h["max_marks"] = m.group(1)

    # Note / Instructions
    if m := re.search(r"Note\s*:\s*([A-Za-z].+?)(?:\n\s*(?:Section|खण्ड|1\.|Q|WQUS)|\n\n|$)", text, re.I | re.DOTALL):
        raw = m.group(1).strip().replace("\n", " ")
        h["note"] = raw[:350]

    return h


def _extract_header(text: str, university: str = "generic") -> Dict[str, str]:
    if university == "mumbai":
        return _extract_header_mumbai(text)
    if university == "abvv":
        return _extract_header_abvv(text)
    return _extract_header_generic(text)


# --------------------------------------------------------------------------
# DEEP HIERARCHICAL STRUCTURE PARSING VIA GROQ LLM (JSON SCHEMA)
# --------------------------------------------------------------------------

STRUCTURE_PROMPT = """You are an expert academic document parser for university exam papers across engineering, sciences, commerce, and humanities.

Given OCR-extracted text from a university exam paper body (AFTER the header), extract ALL questions into a perfectly structured JSON object.

CRITICAL PARSING RULES:

1. HIERARCHY:
   Q1, Q2, Q3... are top-level keys.
   Sub-questions use: i/ii/iii/iv/v (roman) OR a/b/c OR A/B/C (whichever appears).
   MCQ options always use keys A/B/C/D inside "subs" of an MCQ node.

2. BILINGUAL PAPERS (Hindi + English):
   Many papers print Hindi first, then English for the SAME question.
   RULE: Extract ONLY the English text. Discard ALL Hindi/Devanagari text completely.
   Example: ignore "लघु पैमाने की इकाई के निम्न के संदर्भ में मापा" and use only "Small scale unit is measured in terms of:"

3. MCQ DETECTION:
   If choices are labeled (a)/(b)/(c)/(d) or (A)/(B)/(C)/(D): it is an MCQ.
   Set type="mcq". Place each option as a child sub-node with keys A/B/C/D.
   Parent text = English question stem only (before options).
   Each option: { "text": "<English only>", "marks": null, "type": "mcq_option", "subs": {} }
   MCQ options that are bilingual: use ONLY the English option text.

4. ROMAN-NUMERAL SUB-QUESTIONS WITH MCQ OPTIONS:
   Q1 often has sub-questions (i),(ii),(iii),(iv),(v) where each is an MCQ with (a)/(b)/(c)/(d).
   Parse as: Q1.subs = { "i": {type:"mcq", subs:{"A":{...},"B":{...},"C":{...},"D":{...}}}, "ii":{...}, ... }
   Roman numeral keys: "i", "ii", "iii", "iv", "v"
   Option keys: "A", "B", "C", "D"

5. OR / ALTERNATIVES (अथवा/OR):
   When a question is followed by OR (or अथवा/OR), the alternate question uses key "Q<N>_OR".
   Example: Q3 = main, Q3_OR = the OR alternative. Both are siblings at top level.

6. SECTIONS AND UNITS:
   Section-A/Section-B and Unit-I/Unit-II labels are just headers — do NOT create nodes for them.
   All Q-numbered questions go directly as top-level keys.

7. NODE SCHEMA:
   Every node: { "text": "<English only>", "marks": <int or null>, "type": "mcq"|"mcq_option"|"descriptive"|"short", "subs": {} }
   Default type = "descriptive" if unclear.
   Marks from: (5) or [5] or "5 marks" or end of line. Per-question integer only.
   Merge OCR-split lines into one clean sentence.
   Return ONLY raw JSON. No markdown fences, no explanation.

8. SHORT ANSWERS / TOPIC LISTS:
   "Answer any N", "write short note on any N", "explain any N" -> split into child nodes "i","ii"... or "a","b"...
   Parent text = instruction only. Each child = one topic with type="short".

OUTPUT EXAMPLE (SI-003457 style bilingual paper with MCQ + descriptive + OR):
{
  "Q1": {
    "text": "Answer the following objective questions-",
    "marks": 5,
    "type": "mcq",
    "subs": {
      "i": {
        "text": "Small scale unit is measured in terms of:",
        "marks": 1,
        "type": "mcq",
        "subs": {
          "A": { "text": "Capital employed", "marks": null, "type": "mcq_option", "subs": {} },
          "B": { "text": "No. of employees", "marks": null, "type": "mcq_option", "subs": {} },
          "C": { "text": "Quantity of Production", "marks": null, "type": "mcq_option", "subs": {} },
          "D": { "text": "Value of production", "marks": null, "type": "mcq_option", "subs": {} }
        }
      },
      "ii": {
        "text": "Promotion of business is undertaken for following purpose:",
        "marks": 1,
        "type": "mcq",
        "subs": {
          "A": { "text": "Starting a new business", "marks": null, "type": "mcq_option", "subs": {} },
          "B": { "text": "Expansion of existing business", "marks": null, "type": "mcq_option", "subs": {} },
          "C": { "text": "Facing Market Competition", "marks": null, "type": "mcq_option", "subs": {} },
          "D": { "text": "All of the above", "marks": null, "type": "mcq_option", "subs": {} }
        }
      }
    }
  },
  "Q2": {
    "text": "Answer the following questions briefly-",
    "marks": 10,
    "type": "short",
    "subs": {
      "i": { "text": "Name various types of business environment.", "marks": 2, "type": "short", "subs": {} },
      "ii": { "text": "What is the significance of industrial location?", "marks": 2, "type": "short", "subs": {} },
      "iii": { "text": "What is the importance of Business Ethics?", "marks": 2, "type": "short", "subs": {} }
    }
  },
  "Q3": {
    "text": "Explain Scientific Management. How does it help in growth of business?",
    "marks": 5,
    "type": "descriptive",
    "subs": {}
  },
  "Q3_OR": {
    "text": "Illustrate different forms of promotion of business.",
    "marks": 5,
    "type": "descriptive",
    "subs": {}
  },
  "Q5": {
    "text": "What are the different approaches of relationship between business and environment?",
    "marks": 5,
    "type": "descriptive",
    "subs": {}
  },
  "Q5_OR": {
    "text": "Describe the components and types of Environments which impacts the business.",
    "marks": 5,
    "type": "descriptive",
    "subs": {}
  }
}

TEXT TO PARSE:
"""


ABVV_STRUCTURE_PROMPT = """You are an expert parser for ABVV (Atal Bihari Vajpayee Vishwavidyalaya) bilingual exam papers.

This paper has already been cleaned: all Hindi/Devanagari text has been removed. Only English text remains.

NUMBERING CONVENTION OF ABVV PAPERS:
- Top-level questions: "1." "2." "3." ... (plain digits with a dot) → map to Q1, Q2, Q3...
- Sub-questions: "(i)" "(ii)" "(iii)" "(iv)" "(v)" → children of the parent question
- MCQ options: "(a)" "(b)" "(c)" "(d)" → children of the sub-question, type=mcq_option
- OR alternative: line containing only "OR" or "अथवा/OR" → next question uses key "Q<N>_OR"
- Section-A, Section-B, Unit-I, Unit-II → organisational labels, NOT question nodes

CRITICAL RULES:
1. "1." = Q1. The text AFTER "1." on the same line (or the next non-empty line) is Q1's question text.
2. "(i)" under Q1 = Q1's first sub-question → key "i". Its text is on the SAME or NEXT line after "(i)".
3. "(a)" under Q1.i = MCQ option A → {"text":"...","marks":null,"type":"mcq_option","subs":{}}
4. If a question has sub-questions labeled (i)(ii)(iii)..., it is a CONTAINER — recurse into it.
5. If a sub-question has options (a)(b)(c)(d), set type="mcq" and put options in subs as A/B/C/D.
6. MARKS: "1x5=5" means 5 sub-questions each worth 1 mark. "2x5=10" = 5 sub-questions worth 2 marks each.
7. OR alternatives: key pattern is Q<N>_OR. Both Q<N> and Q<N>_OR are siblings at top level.
8. Return ONLY raw JSON. No markdown fences, no explanation.

NODE SCHEMA:
{ "text": "<English only>", "marks": <int or null>, "type": "mcq"|"mcq_option"|"descriptive"|"short", "subs": {} }

--- EXACT EXAMPLE ONLY (DO NOT COPY THIS INTO YOUR OUTPUT) ---
The following is purely to demonstrate how the ABVV numbering scheme maps to the JSON structure.

Example Input text:
  1. Answer the following objective questions-  1x5=5
  (i) Small scale unit is measured in terms of:
  (a) Capital employed  (b) No. of employees  (c) Quantity of Production  (d) Value of production
  (ii) Promotion of business is undertaken for following purpose:
  (a) Starting a new business  (b) Expansion of existing business  (c) Facing Market Competition  (d) All of the above
  (iii) Rationalisation deals with:
  (a) Technological aspect of manufacturing  (b) Elimination of waste  (c) Lesser requirement of Capital  (d) All of the above
  (iv) The main purpose of optimum localisation is:
  (a) Balanced growth  (b) Production and Distribution at lowest cost  (c) Better logistics and transportation  (d) Environmental protection
  (v) Corporate Social responsibility is a:
  (a) Burden to business  (b) Matter of social interest  (c) Matter of interest for both business and society  (d) All of the above
  2. Answer the following questions briefly-  2x5=10
  (i) What do you understand by horizontal and vertical business combinations?
  (ii) Name various types of business environment.
  (iii) What is the significance of industrial location?
  (iv) What is the importance of Business Ethics?
  (v) Explain Scale of Operations.
  Section-B
  Attempt any one question from each unit.  5x4=20
  Unit-I
  3. Explain Scientific Management. How does it help in growth of business?
  OR
  Illustrate different forms of promotion of business.
  Unit-II
  4. What are the different approaches of relationship between business and environment?
  OR
  Describe the components and types of Environments which impacts the business.
  Unit-III
  5. Explain the concept and scope of social responsibility.
  OR
  What is Doctrine of Social Responsibility? Focus on the emerging concepts.
  Unit-IV
  6. Discuss the need and importance of Business Ethics.
  OR
  Differentiate between Business Ethics and Morality.

Expected JSON output:
{
  "Q1": {
    "text": "Answer the following objective questions-",
    "marks": 5,
    "type": "mcq",
    "subs": {
      "i": { "text": "Small scale unit is measured in terms of:", "marks": 1, "type": "mcq",
        "subs": {
          "A": { "text": "Capital employed", "marks": null, "type": "mcq_option", "subs": {} },
          "B": { "text": "No. of employees", "marks": null, "type": "mcq_option", "subs": {} },
          "C": { "text": "Quantity of Production", "marks": null, "type": "mcq_option", "subs": {} },
          "D": { "text": "Value of production", "marks": null, "type": "mcq_option", "subs": {} }
        }
      },
      "ii": { "text": "Promotion of business is undertaken for following purpose:", "marks": 1, "type": "mcq",
        "subs": {
          "A": { "text": "Starting a new business", "marks": null, "type": "mcq_option", "subs": {} },
          "B": { "text": "Expansion of existing business", "marks": null, "type": "mcq_option", "subs": {} },
          "C": { "text": "Facing Market Competition", "marks": null, "type": "mcq_option", "subs": {} },
          "D": { "text": "All of the above", "marks": null, "type": "mcq_option", "subs": {} }
        }
      },
      "iii": { "text": "Rationalisation deals with:", "marks": 1, "type": "mcq",
        "subs": {
          "A": { "text": "Technological aspect of manufacturing", "marks": null, "type": "mcq_option", "subs": {} },
          "B": { "text": "Elimination of waste", "marks": null, "type": "mcq_option", "subs": {} },
          "C": { "text": "Lesser requirement of Capital", "marks": null, "type": "mcq_option", "subs": {} },
          "D": { "text": "All of the above", "marks": null, "type": "mcq_option", "subs": {} }
        }
      },
      "iv": { "text": "The main purpose of optimum localisation is:", "marks": 1, "type": "mcq",
        "subs": {
          "A": { "text": "Balanced growth", "marks": null, "type": "mcq_option", "subs": {} },
          "B": { "text": "Production and Distribution at lowest cost", "marks": null, "type": "mcq_option", "subs": {} },
          "C": { "text": "Better logistics and transportation", "marks": null, "type": "mcq_option", "subs": {} },
          "D": { "text": "Environmental protection", "marks": null, "type": "mcq_option", "subs": {} }
        }
      },
      "v": { "text": "Corporate Social responsibility is a:", "marks": 1, "type": "mcq",
        "subs": {
          "A": { "text": "Burden to business", "marks": null, "type": "mcq_option", "subs": {} },
          "B": { "text": "Matter of social interest", "marks": null, "type": "mcq_option", "subs": {} },
          "C": { "text": "Matter of interest for both business and society", "marks": null, "type": "mcq_option", "subs": {} },
          "D": { "text": "All of the above", "marks": null, "type": "mcq_option", "subs": {} }
        }
      }
    }
  },
  "Q2": {
    "text": "Answer the following questions briefly-",
    "marks": 10,
    "type": "short",
    "subs": {
      "i": { "text": "What do you understand by horizontal and vertical business combinations?", "marks": 2, "type": "short", "subs": {} },
      "ii": { "text": "Name various types of business environment.", "marks": 2, "type": "short", "subs": {} },
      "iii": { "text": "What is the significance of industrial location?", "marks": 2, "type": "short", "subs": {} },
      "iv": { "text": "What is the importance of Business Ethics?", "marks": 2, "type": "short", "subs": {} },
      "v": { "text": "Explain Scale of Operations.", "marks": 2, "type": "short", "subs": {} }
    }
  },
  "Q3": { "text": "Explain Scientific Management. How does it help in growth of business?", "marks": 5, "type": "descriptive", "subs": {} },
  "Q3_OR": { "text": "Illustrate different forms of promotion of business.", "marks": 5, "type": "descriptive", "subs": {} },
  "Q4": { "text": "What are the different approaches of relationship between business and environment?", "marks": 5, "type": "descriptive", "subs": {} },
  "Q4_OR": { "text": "Describe the components and types of Environments which impacts the business.", "marks": 5, "type": "descriptive", "subs": {} },
  "Q5": { "text": "Explain the concept and scope of social responsibility.", "marks": 5, "type": "descriptive", "subs": {} },
  "Q5_OR": { "text": "What is Doctrine of Social Responsibility? Focus on the emerging concepts.", "marks": 5, "type": "descriptive", "subs": {} },
  "Q6": { "text": "Discuss the need and importance of Business Ethics.", "marks": 5, "type": "descriptive", "subs": {} },
  "Q6_OR": { "text": "Differentiate between Business Ethics and Morality.", "marks": 5, "type": "descriptive", "subs": {} }
}

NOW PARSE THE ACTUAL TEXT BELOW. DO NOT OUTPUT ANY QUESTIONS FROM THE EXAMPLE ABOVE. ONLY OUTPUT THE QUESTIONS FOUND IN THIS TEXT:
"""

def _build_tree_via_llm(text: str, university: str = "generic") -> Dict:
    """Use Groq key pool to produce a fully nested JSON question tree."""
    base_prompt = ABVV_STRUCTURE_PROMPT if university == "abvv" else STRUCTURE_PROMPT
    prompt = base_prompt + text[:7000]

    for attempt in range(6):  # try each key once
        llm, api_key = groq_pool.get_llm()
        try:
            resp = llm.invoke([HumanMessage(content=prompt)])
            raw  = resp.content.strip()
            # Strip markdown code fences if present
            raw  = re.sub(r'^```(?:json)?\s*', '', raw)
            raw  = re.sub(r'\s*```$', '', raw)
            tree = json.loads(raw)
            return tree
        except json.JSONDecodeError as e:
            print(f"  [Groq Structure] JSON parse failed (attempt {attempt+1}): {e}")
            return {}  # bad JSON -- don't retry with other keys, same output
        except Exception as e:
            err = str(e)
            if '429' in err or 'rate' in err.lower() or 'token' in err.lower():
                print(f"  [Groq Structure] Key {api_key[:12]}... rate/daily-limited. Trying next key.")
                groq_pool.handle_error(api_key, err)
                continue
            print(f"  [Groq Structure] Error: {e}")
            return {}

    print("  [Groq Structure] All 6 keys exhausted/rate-limited.")
    return {}


def _flatten_tree(tree: Dict, path: List = None, rows: List = None) -> List:
    """Recursively flatten nested tree into tabular rows with 4 hierarchy levels."""
    if path is None:
        path = []
    if rows is None:
        rows = []
    for key, node in tree.items():
        node_type = node.get("type") or "descriptive"
        # Skip MCQ option nodes — their text is included in parent MCQ row
        if node_type == "mcq_option":
            continue

        new_path = path + [key]
        # Pad / truncate to 4 hierarchy columns: Q, Sub, Sub-Sub, Sub-Sub-Sub
        h = (new_path + [""] * 4)[:4]
        marks_val = str(node.get("marks", "") or "")

        # For LEAF MCQ nodes: embed options inline in the text
        text_val  = node.get("text", "")
        subs_map  = node.get("subs") or {}
        is_leaf_mcq = node_type == "mcq" and any(
            sv.get("type") == "mcq_option" for sv in subs_map.values()
        )

        if is_leaf_mcq:
            option_lines = [
                f"  ({ok}) {ov.get('text', '')}"
                for ok, ov in subs_map.items()
                if ov.get("type") == "mcq_option"
            ]
            if option_lines:
                text_val = text_val + "\n" + "\n".join(option_lines)

        rows.append(h + [text_val, marks_val, node_type.upper(), ""])

        # Recurse for container MCQs and all non-MCQ nodes;
        # skip only for leaf MCQs (options already embedded above)
        if subs_map and not is_leaf_mcq:
            _flatten_tree(subs_map, new_path, rows)
    return rows



# --------------------------------------------------------------------------
# BASIC IMAGE PRE-PROCESSING HELPERS
# --------------------------------------------------------------------------

def _preprocess_for_ocr(path: str) -> np.ndarray:
    img = cv2.imread(path)
    if img is None:
        raise FileNotFoundError(f"Cannot read image: {path}")
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # Adaptive threshold for better OCR on varied backgrounds
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                   cv2.THRESH_BINARY, 31, 10)
    return thresh


# Devanagari and extended Hindi Unicode ranges
_HINDI_RE = re.compile(
    r'[\u0900-\u097F'     # Devanagari
    r'\u0966-\u096F'      # Devanagari digits
    r'\uA8E0-\uA8FF'     # Devanagari Extended
    r'\u1CD0-\u1CFF'     # Vedic Extensions
    r']+'
)


def _strip_hindi(text: str) -> str:
    """Remove all Devanagari / Hindi Unicode characters from text."""
    return _HINDI_RE.sub(' ', text)


def _reconstruct_bilingual_lines(text: str) -> str:
    """
    Merge isolated labels like '(i)' or '(a)' with the text on the next line.
    This happens because stripping Hindi leaves the English translation on a new line.
    """
    # Merge (a), (b), (c), (d) or (i), (ii) with the following line
    text = re.sub(r'(?m)^(\s*\([a-zivx]+\))\s*\n\s*', r'\1 ', text)
    # Merge 1., 2. with following line if it's just the number
    text = re.sub(r'(?m)^(\s*\d+\.)\s*\n\s*', r'\1 ', text)
    return text


def _clean_text(text: str) -> str:
    # ── Step 0: Strip ALL Devanagari/Hindi Unicode first ──────────────────
    text = _strip_hindi(text)

    cleaned = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        if "-----" in line or "=====" in line:
            continue
        # After Hindi strip, lines that are mostly non-alphanumeric are garbage
        noise = len(re.findall(r"[^a-zA-Z0-9\s\.\)\(:\-/,;%@]", line))
        if len(line) > 0 and noise > len(line) * 0.40:
            continue
        # Drop lines that are entirely whitespace/punctuation after stripping
        if not re.search(r'[a-zA-Z0-9]', line):
            continue
        cleaned.append(line)
    return "\n".join(cleaned)


# --------------------------------------------------------------------------
# GRAPH NODES
# --------------------------------------------------------------------------

def pdf_ingestion_node(state: QPState) -> QPState:
    task_id = state.get("task_id", "default")
    update_task_progress(task_id, "pdf_ingestion", 10, "Rendering PDF pages with PyMuPDF...")
    try:
        doc = fitz.open(state["pdf_path"])
        image_paths = []
        for i, page in enumerate(doc):
            # 200 DPI for sharper OCR results
            pix = page.get_pixmap(dpi=200)
            
            # If the image is landscape (e.g. a 2-page booklet spread), split it in half
            if pix.width > pix.height * 1.1:
                from PIL import Image
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                w, h = img.size
                left_half = img.crop((0, 0, w//2, h))
                right_half = img.crop((w//2, 0, w, h))
                
                path_l = os.path.join(OUTPUT_DIR, f"{task_id}_page_{i+1}_L.png")
                path_r = os.path.join(OUTPUT_DIR, f"{task_id}_page_{i+1}_R.png")
                left_half.save(path_l)
                right_half.save(path_r)
                image_paths.extend([path_l, path_r])
            else:
                path = os.path.join(OUTPUT_DIR, f"{task_id}_page_{i+1}.png")
                pix.save(path)
                image_paths.append(path)
                
        msg = f"PDF Ingestion complete: {len(image_paths)} page(s) rendered."
        update_task_progress(task_id, "pdf_ingestion", 20, msg)
        return {**state, "image_paths": image_paths,
                "messages": state["messages"] + [AIMessage(content=msg)]}
    except Exception as e:
        err = f"pdf_ingestion failed: {e}"
        update_task_progress(task_id, "pdf_ingestion", 20, f"Error: {err}")
        return {**state, "errors": state["errors"] + [err]}


def ocr_node(state: QPState) -> QPState:
    task_id = state.get("task_id", "default")
    update_task_progress(task_id, "ocr", 30, "Running Tesseract OCR on page images...")
    if not state.get("image_paths"):
        return {**state, "errors": state["errors"] + ["ocr: no image paths"]}

    full_text = ""
    for i, path in enumerate(state["image_paths"]):
        update_task_progress(task_id, "ocr", 30 + int(i * 8 / len(state["image_paths"])),
                             f"Analyzing page {i+1}...")
        thresh = _preprocess_for_ocr(path)
        text = pytesseract.image_to_string(thresh, config=OCR_CONFIG)
        full_text += text + "\n"
        print(f"  Page {i+1}: {len(text)} chars")

    msg = f"OCR complete. Total chars: {len(full_text)}."
    update_task_progress(task_id, "ocr", 40, msg)
    return {**state, "raw_text": full_text,
            "messages": state["messages"] + [AIMessage(content=msg)]}


def cleaning_node(state: QPState) -> QPState:
    task_id = state.get("task_id", "default")
    university = state.get("university", "generic")
    update_task_progress(task_id, "cleaning", 45, "Cleaning OCR noise and artifacts...")
    cleaned = _clean_text(state.get("raw_text", ""))

    if len(cleaned) > 100:
        update_task_progress(task_id, "cleaning", 50, "Groq LLM secondary cleaning pass...")
        is_bilingual = university in ("abvv",)
        bilingual_note = (
            "CRITICAL PRIORITY: This is a BILINGUAL exam paper (Hindi + English).\n"
            "The OCR has scanned BOTH languages, but it lacked Hindi fonts, so Hindi words appear as garbled meaningless ASCII letters (e.g. 'aifeier', 'attat', 'Geren at Seen'). You MUST:\n"
            "  - REMOVE every Hindi word, sentence, or fragment entirely, including all this meaningless garbled ASCII noise.\n"
            "  - Keep ONLY the valid English text for each question.\n"
            "  - IMPORTANT: In bilingual papers, the question number (e.g., '(i)', '1.') is often only written ONCE before the Hindi text. If you delete a garbled Hindi question, YOU MUST PRESERVE its question number and attach it to the English translation!\n"
            "  - IMPORTANT: If a question has duplicate multiple-choice options (one set for Hindi, one set for English like '(a) 9 (b) 18...' appearing twice), KEEP ONLY ONE SET of options.\n"
        ) if is_bilingual else (
            "If any text is written in Hindi (Devanagari script) or garbled ASCII noise, COMPLETELY REMOVE it. Only retain valid English text.\n"
        )
        prompt = (
            "You are an OCR text cleaning assistant for university exam papers.\n"
            + bilingual_note +
            "Additional rules:\n"
            "1. Fix broken words and merge lines split due to two-column layout.\n"
            "2. Remove page footers, watermarks, header repetitions (like 'SI-003457', 'Turn Over', 'Continued'). Note that the main top header has already been extracted, you can strip it.\n"
            "3. Preserve question numbers (1., 2., 3... or Q1, Q2...), sub-question labels (i, ii, iii or a, b, c), MCQ option labels ((a)(b)(c)(d) or (A)(B)(C)(D)), and marks notation.\n"
            "4. Preserve Section-A / Section-B / Unit-I / Unit-II headings.\n"
            "5. Do NOT reorder or rephrase questions. Output the cleaned text in the exact same order.\n"
            "Return ONLY the cleaned English text. No preamble or explanations.\n\n"
            f"TEXT:\n{cleaned[:7000]}"
        )
        for attempt in range(6):
            llm, api_key = groq_pool.get_llm(model_name="meta-llama/llama-4-scout-17b-16e-instruct")
            try:
                resp    = llm.invoke([HumanMessage(content=prompt)])
                cleaned = resp.content.strip()
                update_task_progress(task_id, "cleaning", 55, "Groq LLM cleaning complete.")
                break
            except Exception as e:
                err = str(e)
                if '429' in err or 'rate' in err.lower() or 'token' in err.lower():
                    print(f"  [Cleaning] Key {api_key[:12]}... rate/daily-limited. Trying next.")
                    groq_pool.handle_error(api_key, err)
                    continue
                update_task_progress(task_id, "cleaning", 55, f"LLM cleaning skipped: {e}")
                break
        else:
            update_task_progress(task_id, "cleaning", 55, "LLM cleaning skipped -- all keys rate-limited.")

    update_task_progress(task_id, "cleaning", 60, f"Cleaning done. {len(cleaned)} chars retained.")
    return {**state, "clean_text": cleaned,
            "messages": state["messages"] + [AIMessage(content=f"Cleaning done. {len(cleaned)} chars.")]}


def header_extraction_node(state: QPState) -> QPState:
    task_id    = state.get("task_id", "default")
    university = state.get("university", "generic")
    update_task_progress(task_id, "header_extraction", 65,
                         f"Extracting header ({university.upper()} mode)...")

    # Extract header from the raw OCR text (before LLM cleaning) because
    # the LLM often aggressively deletes the header block thinking it's a watermark.
    raw_text = state.get("raw_text", state.get("clean_text", ""))
    header = _extract_header(raw_text, university)

    # Separate body text using the LLM-cleaned text so we don't parse garbage
    clean_text = state.get("clean_text", "")

    # Handle both: 'Q1' / 'Q.1' style (Mumbai) and '1.' plain number style (ABVV)
    # Try Q-style first, then plain number style.
    body = None
    split = re.split(r"(?:^|\n)\s*(Q\.?\s*1\b)", clean_text, maxsplit=1, flags=re.I | re.M)
    if len(split) > 1:
        body = (split[1] + split[2]).strip() if len(split) > 2 else split[1].strip()
    else:
        # ABVV / plain numbered: split on first standalone '1.' or '1 .'
        split2 = re.split(r"(?:^|\n)(\s*1\s*\.(?:\s|$))", clean_text, maxsplit=1, flags=re.M)
        if len(split2) > 1:
            body = (split2[1] + split2[2]).strip() if len(split2) > 2 else split2[1].strip()
        else:
            body = clean_text  # fallback: use entire clean text as body

    update_task_progress(task_id, "header_extraction", 70,
                         f"Header extracted. Subject: {header.get('subject') or 'Unknown'}")
    return {**state, "header": header, "body_text": body,
            "messages": state["messages"] + [AIMessage(content=f"Header: {header}")]}


def normalization_node(state: QPState) -> QPState:
    task_id    = state.get("task_id", "default")
    university = state.get("university", "generic")
    update_task_progress(task_id, "normalization", 72,
                         "Normalizing structural boundaries for hierarchical parsing...")
    text = state.get("body_text", "")

    # Final Hindi strip pass in case LLM cleaning left residue
    text = _strip_hindi(text)

    # For bilingual (ABVV) papers: reconstruct split lines so the LLM sees
    # '(i) Small scale unit...' not '(i)\nSmall scale unit...' on separate lines
    if university == "abvv":
        text = _reconstruct_bilingual_lines(text)

    # Ensure Q-style markers have a preceding newline
    text = re.sub(r"(?<!\n)(Q\.?\s*\d+)", r"\n\1", text, flags=re.I)
    # Ensure plain number markers '1.' '2.' etc start on their own line
    text = re.sub(r"(?<!\n)(^\s*\d+\.\s)", r"\n\1", text, flags=re.I | re.M)
    # Remove page artefacts like 'Turn Over', 'Continued', 'SI-003457' repeats
    text = re.sub(r'\(Turn\s*Over\)', '', text, flags=re.I)
    text = re.sub(r'\(Continued\)', '', text, flags=re.I)
    text = re.sub(r'\bSI[-\s]?\d{4,8}\b', '', text, flags=re.I)
    # Remove standalone page numbers like (2), (3), (4) on their own line
    text = re.sub(r'^\(\d+\)\s*$', '', text, flags=re.M)
    # Collapse 3+ blank lines into 1
    text = re.sub(r'\n{3,}', '\n\n', text)

    update_task_progress(task_id, "normalization", 78, "Normalization complete.")
    return {**state, "normalized": text,
            "messages": state["messages"] + [AIMessage(content="Normalization done.")]}


def structure_node(state: QPState) -> QPState:
    task_id    = state.get("task_id", "default")
    university = state.get("university", "generic")
    update_task_progress(task_id, "structure", 80,
                         "Invoking Groq Llama-3.3-70b for deep hierarchical JSON tree construction...")

    normalized = state.get("normalized", "")

    tree = _build_tree_via_llm(normalized, university)
    q_count = len(tree)
    update_task_progress(task_id, "structure", 88,
                         f"JSON tree built: {q_count} main question(s) extracted.")
    print(f"  {q_count} top-level question(s) found via Groq JSON parser.")

    return {**state, "structured": tree,
            "messages": state["messages"] + [AIMessage(content=f"Structure: {q_count} main questions.")]}


def _find_diagram_regions(img: np.ndarray) -> List[tuple]:
    """Use OpenCV to detect table grids and circuit-like regions precisely."""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape

    # ── 1. Detect table grid: morphological line detection ──
    _, thresh = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
    kernel_h = cv2.getStructuringElement(cv2.MORPH_RECT, (max(1, w // 8), 1))
    kernel_v = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(1, h // 8)))
    h_lines  = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel_h)
    v_lines  = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel_v)
    grid     = cv2.add(h_lines, v_lines)

    # ── 2. Detect circuit/diagram: Canny edge density ──
    edges = cv2.Canny(gray, 50, 150)
    combined = cv2.add(grid, edges)

    # Dilate to merge nearby blobs
    kernel_d = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 25))
    dilated  = cv2.dilate(combined, kernel_d, iterations=3)

    contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    regions = []
    min_area = h * w * 0.018  # at least 1.8% of page area
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < min_area:
            continue
        x, y, cw, ch = cv2.boundingRect(cnt)
        # Skip full-page blobs (just noise)
        if cw > w * 0.92 and ch > h * 0.92:
            continue
        # Pad by 12 pixels
        x1 = max(0, x - 12)
        y1 = max(0, y - 12)
        x2 = min(w, x + cw + 12)
        y2 = min(h, y + ch + 12)
        regions.append((x1, y1, x2, y2))
    return regions


def diagram_node(state: QPState) -> QPState:
    task_id     = state.get("task_id", "default")
    image_paths = state.get("image_paths", [])
    structured  = state.get("structured", {})
    update_task_progress(task_id, "diagram", 90, "Scanning for diagram/figure keywords...")

    keywords = {
        "diagram", "figure", "draw", "sketch", "plot", "graph", "circuit",
        "flowchart", "network", "gate", "schematic", "truth table", "waveform",
        "table", "block diagram", "circuit diagram", "karnaugh", "k-map",
        "timing diagram", "state machine", "ladder"
    }
    paths: List[str] = []

    def _has_kw(tree: Dict) -> bool:
        for node in tree.values():
            txt = node.get("text", "").lower()
            if any(kw in txt for kw in keywords):
                return True
            if node.get("subs") and _has_kw(node["subs"]):
                return True
        return False

    if image_paths and _has_kw(structured):
        for page_idx, path in enumerate(image_paths):
            img = cv2.imread(path)
            if img is None:
                continue
            regions = _find_diagram_regions(img)
            if regions:
                for reg_idx, (x1, y1, x2, y2) in enumerate(regions):
                    crop     = img[y1:y2, x1:x2]
                    out_path = os.path.join(OUTPUT_DIR, f"{task_id}_diagram_p{page_idx+1}_r{reg_idx+1}.png")
                    cv2.imwrite(out_path, crop)
                    paths.append(out_path)
        msg = f"Precise diagram/table region(s) detected and cropped: {len(paths)} image(s)."
    else:
        msg = "No diagram keywords detected — skipping visual extraction."

    update_task_progress(task_id, "diagram", 92, msg)
    return {**state, "diagram_paths": paths,
            "messages": state["messages"] + [AIMessage(content=msg)]}


def flatten_node(state: QPState) -> QPState:
    task_id = state.get("task_id", "default")
    update_task_progress(task_id, "flatten", 94, "Flattening nested tree into tabular rows...")
    rows = _flatten_tree(state.get("structured", {}))
    update_task_progress(task_id, "flatten", 96, f"Generated {len(rows)} total rows.")
    return {**state, "flat_rows": rows,
            "messages": state["messages"] + [AIMessage(content=f"Flattened: {len(rows)} rows.")]}


def excel_writer_node(state: QPState) -> QPState:
    task_id    = state.get("task_id", "default")
    university = state.get("university", "generic")
    update_task_progress(task_id, "excel_writer", 97, "Generating Excel workbook...")

    header   = state.get("header", {})
    rows     = state.get("flat_rows", [])
    diagrams = state.get("diagram_paths", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Question Paper Extraction"

    # ── Title row ──────────────────────────────────────────────
    ws.append(["EXTRACTA — AI Question Paper Extraction Report"])
    ws.append([])

    # ── Header metadata block ──────────────────────────────────
    metadata_fields = [
        ("Paper Code",    "paper_code"),
        ("Subject",       "subject"),
        ("Date",          "date"),
        ("Time",          "time"),
        ("Max Marks",     "max_marks"),
        ("QP Code",       "qp_code"),
    ]
    if university == "mumbai":
        metadata_fields += [
            ("Branch",    "branch"),
            ("Semester",  "semester"),
            ("Scheme",    "scheme"),
            ("Note",      "note"),
        ]
    elif university == "abvv":
        metadata_fields += [
            ("Semester",      "semester"),
            ("Exam Session",  "exam_session"),
            ("Exam Type",     "exam"),
            ("Paper Type",    "paper_type"),
            ("Note",          "note"),
        ]
    else:
        metadata_fields += [("Exam", "exam")]


    for label, key in metadata_fields:
        ws.append([label, header.get(key, "")])

    ws.append([])

    # ── Column headers ─────────────────────────────────────────
    ws.append(["Q", "Sub", "Sub-Sub", "Sub-Sub-Sub", "Question Text", "Marks", "Type", "Visual"])
    start_row = ws.max_row + 1

    for row in rows:
        while len(row) < 8:
            row.append("")
        ws.append(row)

    # ── Embed diagrams ─────────────────────────────────────────
    for i, path in enumerate(diagrams):
        if os.path.exists(path):
            try:
                ws.add_image(ExcelImage(path), f"H{start_row + i}")
            except Exception as e:
                print(f"  Image embed error: {e}")

    # ── Column widths ──────────────────────────────────────────
    for col, width in [("A",8),("B",8),("C",8),("D",8),("E",90),("F",8),("G",12),("H",30)]:
        ws.column_dimensions[col].width = width

    pdf_path = state.get("pdf_path", "")
    if pdf_path:
        base_name = os.path.basename(pdf_path)
        name_no_ext, _ = os.path.splitext(base_name)
        out_filename = f"sheet_{name_no_ext}.xlsx"
    else:
        out_filename = f"sheet_{task_id}.xlsx"

    out_path = os.path.join(OUTPUT_DIR, out_filename)
    wb.save(out_path)
    update_task_progress(task_id, "excel_writer", 98, f"Workbook saved: {out_filename}")
    return {**state, "excel_path": out_path,
            "messages": state["messages"] + [AIMessage(content=f"Excel: {out_path}")]}


def validation_node(state: QPState) -> QPState:
    task_id = state.get("task_id", "default")
    update_task_progress(task_id, "validation", 99, "Validating extraction results...")

    ok    = bool(state.get("flat_rows")) and os.path.exists(state.get("excel_path", ""))
    retry = state.get("retry_count", 0)

    if ok:
        msg = f"Validation PASSED — {len(state['flat_rows'])} rows extracted successfully."
    else:
        msg = f"Validation FAILED — will retry (attempt {retry+1})."

    summary = ""
    if LLM_AVAILABLE and ok:
        try:
            log = "\n".join(m.content for m in state["messages"] if hasattr(m, "content"))
            resp = LLM.invoke([HumanMessage(
                content="Summarise this exam extraction in 2 concise sentences covering subject, question count, and status:\n\n" + log[-2000:]
            )])
            summary = resp.content
        except Exception:
            pass

    update_task_progress(task_id, "validation", 100, msg)
    return {**state,
            "validation_ok": ok,
            "retry_count":   retry + (0 if ok else 1),
            "messages": state["messages"] + [AIMessage(content=summary or msg)]}


# --------------------------------------------------------------------------
# CONDITIONAL EDGE
# --------------------------------------------------------------------------

def retry_router(state: QPState) -> str:
    if state.get("validation_ok"):
        return "end"
    if state.get("retry_count", 0) < MAX_RETRIES:
        return "retry"
    return "end"


# --------------------------------------------------------------------------
# GRAPH ASSEMBLY
# --------------------------------------------------------------------------

def build_graph() -> StateGraph:
    g = StateGraph(QPState)
    for name, fn in [
        ("pdf_ingestion",     pdf_ingestion_node),
        ("ocr",               ocr_node),
        ("cleaning",          cleaning_node),
        ("header_extraction", header_extraction_node),
        ("normalization",     normalization_node),
        ("structure",         structure_node),
        ("diagram",           diagram_node),
        ("flatten",           flatten_node),
        ("excel_writer",      excel_writer_node),
        ("validation",        validation_node),
    ]:
        g.add_node(name, fn)

    g.set_entry_point("pdf_ingestion")
    g.add_edge("pdf_ingestion",     "ocr")
    g.add_edge("ocr",               "cleaning")
    g.add_edge("cleaning",          "header_extraction")
    g.add_edge("header_extraction", "normalization")
    g.add_edge("normalization",     "structure")
    g.add_edge("structure",         "diagram")
    g.add_edge("diagram",           "flatten")
    g.add_edge("flatten",           "excel_writer")
    g.add_edge("excel_writer",      "validation")
    g.add_conditional_edges("validation", retry_router, {"retry": "ocr", "end": END})
    return g


# --------------------------------------------------------------------------
# PUBLIC ENTRY POINT
# --------------------------------------------------------------------------

def run_agent_pipeline(pdf_path: str, task_id: str, university: str = "generic") -> QPState:
    memory = MemorySaver()
    graph  = build_graph().compile(checkpointer=memory)

    initial: QPState = {
        "task_id":       task_id,
        "university":    university,
        "pdf_path":      pdf_path,
        "image_paths":   [],
        "raw_text":      "",
        "clean_text":    "",
        "header":        {},
        "body_text":     "",
        "normalized":    "",
        "structured":    {},
        "diagram_paths": [],
        "flat_rows":     [],
        "excel_path":    "",
        "validation_ok": False,
        "retry_count":   0,
        "messages":      [HumanMessage(content=f"Process: {pdf_path} [{university}]")],
        "errors":        [],
    }
    config = {"configurable": {"thread_id": f"qp-{task_id}"}}
    return graph.invoke(initial, config=config)


# ==========================================================================
# ── EXTRACTA V2.2 UPGRADE: HOLISTIC MODEL ANSWER SHEET AGENT ──────────────
# ==========================================================================

class GroqKeysPool:
    """Rotates through ext1-ext6 Groq keys loaded from .env."""
    def __init__(self):
        # Load ext1-ext6 from environment (populated by .env via dotenv)
        self.keys = [
            k for k in [
                os.environ.get("GROQ_EXT1"),
                os.environ.get("GROQ_EXT2"),
                os.environ.get("GROQ_EXT3"),
                os.environ.get("GROQ_EXT4"),
                os.environ.get("GROQ_EXT5"),
                os.environ.get("GROQ_EXT6"),
            ]
            if k  # skip any None / empty entries
        ]
        if not self.keys:
            print("[GroqKeysPool] WARNING: No ext keys found in .env. Add GROQ_EXT1..EXT6.")
        else:
            print(f"[GroqKeysPool] Loaded {len(self.keys)} extraction key(s) from .env.")
        self._index   = 0
        self._lock    = threading.Lock()
        self.cooldowns = {k: 0.0 for k in self.keys}

    def _parse_retry_seconds(self, err_msg: str) -> float:
        """Parse 'try again in Xm Y.Zs' from Groq error messages."""
        m = re.search(r'try again in (\d+)m([\d.]+)s', err_msg)
        if m:
            return int(m.group(1)) * 60 + float(m.group(2)) + 5
        m = re.search(r'try again in ([\d.]+)s', err_msg)
        if m:
            return float(m.group(1)) + 5
        return 90.0  # safe default

    def handle_error(self, key: str, err_msg: str):
        """Intelligently set cooldown based on error type."""
        with self._lock:
            if key not in self.cooldowns:
                return
            # Daily token limit (TPD) — exhausted for the day
            if 'tokens per day' in err_msg or 'TPD' in err_msg:
                cooldown = 23 * 3600  # 23 hours
                print(f"[GroqKeysPool] Key {key[:12]}... DAILY LIMIT exhausted. Cooling 23h.")
            # Per-minute or per-request rate limit
            elif '429' in err_msg or 'rate' in err_msg.lower():
                cooldown = self._parse_retry_seconds(err_msg)
                print(f"[GroqKeysPool] Key {key[:12]}... rate-limited. Cooling {cooldown:.0f}s.")
            else:
                cooldown = 30.0
                print(f"[GroqKeysPool] Key {key[:12]}... error. Cooling 30s.")
            self.cooldowns[key] = time.time() + cooldown

    def get_llm(self, model_name="llama-3.3-70b-versatile"):
        with self._lock:
            now = time.time()
            for _ in range(len(self.keys)):
                key = self.keys[self._index]
                self._index = (self._index + 1) % len(self.keys)
                if now >= self.cooldowns[key]:
                    return ChatGroq(model=model_name,
                                    temperature=0.1, api_key=key), key
            # All keys on cooldown — use least-expired one
            best_key = min(self.keys, key=lambda k: self.cooldowns[k])
            return ChatGroq(model=model_name,
                            temperature=0.1, api_key=best_key), best_key

groq_pool = GroqKeysPool()


def classify_question(text: str, q_type: str = "descriptive") -> str:
    """Classify question for answer generation. Respects pre-tagged MCQ type from LLM parser."""
    if q_type == "mcq":
        return "MCQ"
    t = text.lower()
    if any(w in t for w in ["differentiate", "compare", "distinguish", "difference"]):
        return "Differences"
    elif any(w in t for w in ["define", "what is", "explain", "state", "list", "write short note"]):
        return "Define"
    elif any(w in t for w in ["justify", "prove", "why", "compulsory"]):
        return "Justify"
    elif any(w in t for w in ["analyze", "solve", "compute", "calculate", "design", "diagram", "draw"]):
        return "Analysis"
    return "General Explanation"


ANSWERS_PROMPTS = {
    "MCQ": (
        "You are an expert university examiner. A multiple-choice question is given below with its options.\n"
        "Question Stem: '{question}'\n"
        "Options:\n{options}\n\n"
        "STRICT OUTPUT FORMAT (PLAIN TEXT ONLY -- NO ** OR * SYMBOLS ANYWHERE):\n"
        "1. === CORRECT ANSWER ===\n"
        "   State: The correct answer is option [LETTER]: [Option Text]\n"
        "2. === EXPLANATION ===\n"
        "   Provide a clear, technically precise explanation in {num_points} sentences justifying WHY this option is correct.\n"
        "3. === WHY OTHER OPTIONS ARE WRONG ===\n"
        "   For each incorrect option, write one sentence explaining why it is wrong.\n"
        "4. PLAIN TEXT ONLY. No asterisks, no bold markers.\n"
        "5. Tone: strictly academic, formal, zero emojis."
    ),
    "Differences": (
        "You are a university professor writing a model answer for a comparison/differentiation question.\n"
        "Question: '{question}'\n"
        "The two things being compared are: TERM_1 = '{concept1}'  and  TERM_2 = '{concept2}'\n"
        "Total Marks: {marks}\n\n"
        "YOUR TASK: Write a complete, detailed university-level comparative answer. Follow this EXACT structure:\n"
        "\n"
        "=== OVERVIEW ===\n"
        "Write 2-3 sentences introducing both '{concept1}' and '{concept2}' and why comparing them is meaningful.\n"
        "\n"
        "=== COMPARISON TABLE ===\n"
        "Create a markdown table with EXACTLY 3 columns: Parameter | {concept1} | {concept2}\n"
        "- Include EXACTLY {num_points} rows (one per distinct comparison parameter).\n"
        "- BOTH the '{concept1}' column AND the '{concept2}' column MUST have a value in EVERY single row. Do NOT leave any cell blank or empty.\n"
        "- Each cell: max 10 words, precise, technical.\n"
        "- Parameters to cover (use these or similar technical ones): Definition, Scope, Application, Enforcement/Source, Focus/Goal, Nature, Example.\n"
        "- Format: | Parameter | value for {concept1} | value for {concept2} |\n"
        "- Start the table with the header row: | Parameter | {concept1} | {concept2} |\n"
        "\n"
        "=== {concept1}: DETAILED EXPLANATION ===\n"
        "Write 3-4 sentences: definition, principle, real-world use case.\n"
        "\n"
        "=== {concept2}: DETAILED EXPLANATION ===\n"
        "Write 3-4 sentences: definition, principle, real-world use case.\n"
        "\n"
        "RULES:\n"
        "- PLAIN TEXT ONLY. No ** or * markdown bold/italic.\n"
        "- The comparison table MUST have values in EVERY cell — never leave a column empty.\n"
        "- Tone: strictly academic, formal, technically precise, zero emojis."
    ),
    "Define": (
        "Write a highly descriptive, super-technical university-style answer for:\n"
        "Question: '{question}'\n"
        "Total Marks: {marks}\n\n"
        "STRICT OUTPUT FORMAT (PLAIN TEXT ONLY -- NO ** OR * SYMBOLS ANYWHERE):\n"
        "1. Write a precise formal academic definition (2-3 sentences).\n"
        "2. === THEORETICAL BACKGROUND ===\n"
        "   Provide the full theoretical framework as {num_points} numbered points.\n"
        "3. === KEY CHARACTERISTICS ===\n"
        "   List {num_points} core attributes as numbered points.\n"
        "4. === APPLICATIONS ===\n"
        "   List 2-3 real-world engineering applications as numbered points.\n"
        "5. Write formulas in plain notation only: F = ma, v = u + at.\n"
        "6. PLAIN TEXT ONLY. No asterisks. UPPERCASE for emphasis only.\n"
        "7. Tone: strictly academic, formal, super-technical, zero emojis."
    ),
    "Justify": (
        "Write a highly technical, university-style justification answer for:\n"
        "Question: '{question}'\n"
        "Total Marks: {marks}\n\n"
        "STRICT OUTPUT FORMAT (PLAIN TEXT ONLY -- NO ** OR * SYMBOLS ANYWHERE):\n"
        "1. === THESIS STATEMENT ===\n"
        "   State the thesis in 1 clear sentence.\n"
        "2. === JUSTIFICATION ===\n"
        "   Provide a {num_points}-step logical argumentation.\n"
        "   Number each step. Each must be formally written and technically precise.\n"
        "3. === SUPPORTING EVIDENCE ===\n"
        "   Include formulas in plain notation, real-world evidence, or derived results.\n"
        "4. === CONCLUSION ===\n"
        "   Summarize in 2 sentences.\n"
        "5. PLAIN TEXT ONLY. No asterisks.\n"
        "6. Tone: strictly academic, formal, super-technical, zero emojis."
    ),
    "Analysis": (
        "Write a comprehensive, highly technical university-style engineering/mathematical analysis for:\n"
        "Question: '{question}'\n"
        "Total Marks: {marks}\n\n"
        "STRICT OUTPUT FORMAT (PLAIN TEXT ONLY -- NO ** OR * SYMBOLS ANYWHERE):\n"
        "1. === PROBLEM STATEMENT ===\n"
        "   Restate the problem formally in 2 sentences.\n"
        "2. === APPROACH / METHODOLOGY ===\n"
        "   Describe the method, tools, or algorithm to be used.\n"
        "3. === STEP-BY-STEP SOLUTION ===\n"
        "   Solve or design in exactly {num_points} clearly numbered steps.\n"
        "   Show all formulas in plain notation: Z = V/I, P = VI cos(phi).\n"
        "4. === RESULT / CONCLUSION ===\n"
        "   State the final result and its engineering significance.\n"
        "5. PLAIN TEXT ONLY. No asterisks, no dollar signs.\n"
        "6. Tone: strictly academic, formal, super-technical, zero emojis."
    ),
    "General Explanation": (
        "Write a formal, super-technical, highly descriptive university-style explanation for:\n"
        "Question: '{question}'\n"
        "Total Marks: {marks}\n\n"
        "STRICT OUTPUT FORMAT (PLAIN TEXT ONLY -- NO ** OR * SYMBOLS ANYWHERE):\n"
        "1. === OVERVIEW AND THEORY ===\n"
        "   Provide a 2-3 sentence formal overview.\n"
        "2. === TECHNICAL DEEP-DIVE ===\n"
        "   Provide {num_points} numbered technical points. Each must be a complete sentence.\n"
        "3. === ACADEMIC SYNTHESIS ===\n"
        "   Synthesize in 3-4 sentences covering significance, limitations, and real-world relevance.\n"
        "4. PLAIN TEXT ONLY. No asterisks. UPPERCASE for emphasis.\n"
        "5. Write formulas in plain notation: E = mc^2, F = G*m1*m2/r^2.\n"
        "6. Tone: strictly academic, formal, super-technical, zero emojis."
    )
}



def _marks_to_points(marks) -> int:
    """Return number of answer points appropriate for given marks."""
    try:
        m = int(marks or 0)
    except (ValueError, TypeError):
        m = 0
    if m <= 0:   return 6
    if m <= 3:   return 4
    if m <= 5:   return 7
    if m <= 7:   return 9
    if m <= 10:  return 12
    return 14


def _extract_comparison_concepts(text: str):
    """
    Robustly extract the two terms being compared from a differentiation question.
    Handles patterns: 'between X and Y', 'X and Y', 'X vs Y', 'X versus Y',
    'differentiate X and Y', 'compare X with Y', etc.
    Returns (concept1, concept2).
    """
    t = text.strip()

    # Pattern 1: 'between X and Y' (most reliable)
    m = re.search(
        r'between\s+([A-Za-z][A-Za-z\s\-\']+?)\s+and\s+([A-Za-z][A-Za-z\s\-\']+?)'
        r'(?:\s*[.?!,]|$)',
        t, re.I
    )
    if m:
        return m.group(1).strip().title(), m.group(2).strip().title()

    # Pattern 2: 'X and Y' after a verb keyword
    m = re.search(
        r'(?:differentiate|compare|distinguish|contrast|difference\s+between)\s+'
        r'([A-Za-z][A-Za-z\s\-\']+?)\s+(?:and|vs\.?|versus|with)\s+'
        r'([A-Za-z][A-Za-z\s\-\']+?)(?:\s*[.?!,]|$)',
        t, re.I
    )
    if m:
        return m.group(1).strip().title(), m.group(2).strip().title()

    # Pattern 3: 'X vs Y' anywhere
    m = re.search(
        r'([A-Za-z][A-Za-z\s\-\']+?)\s+(?:vs\.?|versus)\s+'
        r'([A-Za-z][A-Za-z\s\-\']+?)(?:\s*[.?!,]|$)',
        t, re.I
    )
    if m:
        return m.group(1).strip().title(), m.group(2).strip().title()

    # Fallback: extract two longest noun phrases after the first keyword
    after_kw = re.split(r'\b(?:differentiate|compare|distinguish|explain|describe)\b', t, maxsplit=1, flags=re.I)
    search_text = after_kw[1] if len(after_kw) > 1 else t
    # Find anything separated by 'and'
    m = re.search(r'([A-Za-z][A-Za-z\s\-\']{2,30?})\s+and\s+([A-Za-z][A-Za-z\s\-\']{2,30?})', search_text, re.I)
    if m:
        return m.group(1).strip().title(), m.group(2).strip().title()

    # Last resort: just label them generically using key nouns
    nouns = re.findall(r'[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*', t)
    c1 = nouns[0] if len(nouns) > 0 else "Concept A"
    c2 = nouns[1] if len(nouns) > 1 else "Concept B"
    return c1, c2


def solve_single_question(item: Dict[str, Any]) -> Dict[str, Any]:
    key      = item["key"]
    text     = item["text"]
    marks    = item.get("marks", 0)
    q_type   = item.get("q_type", "descriptive")
    options  = item.get("options", [])  # list of dicts {letter, text} for MCQ
    category = classify_question(text, q_type)
    num_pts  = _marks_to_points(marks)

    # Robustly extract the two concepts for Differences prompts
    concept1, concept2 = _extract_comparison_concepts(text)

    if category == "MCQ":
        options_str = "\n".join(
            f"  ({o['letter']}) {o['text']}" for o in options
        ) if options else "Options not available"
        prompt = ANSWERS_PROMPTS["MCQ"].format(
            question   = text,
            options    = options_str,
            num_points = max(num_pts, 3),
        )
    else:
        prompt = ANSWERS_PROMPTS[category].format(
            question   = text,
            marks      = marks or "Not specified",
            num_points = num_pts,
            concept1   = concept1,
            concept2   = concept2,
        )

    for attempt in range(5):
        llm, api_key = groq_pool.get_llm(model_name="meta-llama/llama-4-scout-17b-16e-instruct")
        try:
            resp = llm.invoke([HumanMessage(content=prompt)])
            return {
                "key":      key,
                "text":     text,
                "marks":    marks,
                "category": category,
                "q_type":   q_type,
                "options":  options,
                "answer":   resp.content.strip()
            }
        except Exception as e:
            err_msg = str(e)
            if "429" in err_msg or "rate" in err_msg.lower() or "token" in err_msg.lower():
                groq_pool.handle_error(api_key, err_msg)
                continue
            else:
                time.sleep(2)

    return {
        "key":      key,
        "text":     text,
        "marks":    marks,
        "category": category,
        "q_type":   q_type,
        "options":  options,
        "answer":   "Model Answer could not be generated due to service limits."
    }


def _answer_sort_key(key: str) -> list:
    """
    Natural hierarchical sort key for answer keys like:
    Q1, Q1.i, Q1.ii, Q2, Q3, Q3_OR, Q10, Q10.i
    Sorts: Q1 < Q1.i < Q1.ii < Q2 < Q3 < Q3_OR < Q10 < Q10.i
    """
    roman_map = {"i":1,"ii":2,"iii":3,"iv":4,"v":5,
                 "vi":6,"vii":7,"viii":8,"ix":9,"x":10}
    parts = key.split(".")
    result = []
    for part in parts:
        is_or = 0
        core  = part
        if "_OR" in part.upper():
            is_or = 1
            core  = re.sub(r'_OR', '', part, flags=re.I)
        m = re.match(r'[Qq](\d+)', core)
        if m:
            result.extend([int(m.group(1)), is_or])
        else:
            low = core.lower()
            if low in roman_map:
                result.extend([roman_map[low], is_or])
            elif len(low) == 1 and low.isalpha():
                result.extend([ord(low) - ord('a') + 1, is_or])
            else:
                result.extend([0, is_or])
    return result


def generate_answers_for_tree(structured_tree: Dict, task_id: str) -> List[Dict[str, Any]]:

    flat_questions = []

    def traverse(node_map: Dict, path: str = "", parent_type: str = "descriptive"):
        for k, v in node_map.items():
            current_key = f"{path}.{k}" if path else k
            node_type = v.get("type", parent_type)

            # Skip MCQ option child nodes — their parent MCQ node handles them
            if node_type == "mcq_option":
                continue

            if v.get("text"):
                item = {
                    "key":    current_key,
                    "text":   v["text"],
                    "marks":  v.get("marks", 0),
                    "q_type": node_type,
                }
                # If this is a LEAF MCQ (has direct mcq_option children), collect its options
                subs = v.get("subs") or {}
                has_direct_options = any(
                    sv.get("type") == "mcq_option" for sv in subs.values()
                )
                if node_type == "mcq" and has_direct_options:
                    item["options"] = [
                        {"letter": opt_key, "text": opt_val.get("text", "")}
                        for opt_key, opt_val in subs.items()
                        if opt_val.get("type") == "mcq_option"
                    ]
                flat_questions.append(item)

            # Recurse into subs:
            # - Always recurse for non-MCQ nodes
            # - Recurse for CONTAINER MCQs (sub-questions like i, ii, iii)
            # - SKIP only for LEAF MCQs (direct A/B/C/D options already captured)
            subs_map = v.get("subs") or {}
            is_leaf_mcq = any(sv.get("type") == "mcq_option" for sv in subs_map.values())
            if subs_map and not is_leaf_mcq:
                traverse(subs_map, current_key, node_type)

    traverse(structured_tree)

    if not flat_questions:
        return []

    print(f"[{task_id}] Solving {len(flat_questions)} questions in parallel with 6 Groq Keys...")

    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(solve_single_question, item): item for item in flat_questions}
        for index, future in enumerate(concurrent.futures.as_completed(futures)):
            res = future.result()
            results.append(res)
            percentage = 10 + int((index + 1) * 80 / len(flat_questions))
            update_task_progress(task_id, "answering", percentage,
                                 f"Solved {index+1} of {len(flat_questions)}: Question {res['key']}")

    results.sort(key=lambda x: _answer_sort_key(x["key"]))
    return results




# ── FPDF2 PDF Compiler ──

def sanitize_for_pdf(text: str) -> str:
    replacements = {
        "\u2014": " -- ", # em dash
        "\u2013": " - ",  # en dash
        "\u201c": '"',    # smart opening double quote
        "\u201d": '"',    # smart closing double quote
        "\u2018": "'",    # smart opening single quote
        "\u2019": "'",    # smart closing single quote
        "\u2208": " in ",
        "\u222b": "Integral ",
        "\u03b8": "theta",
        "\u03c0": "pi",
        "\u03bb": "lambda",
        "\u03b4": "delta",
        "\u2211": "Sum",
        "\u2192": " -> ",
        "\u2260": " != ",
        "\u2264": " <= ",
        "\u2265": " >= ",
        "\u00b1": " +/- ",
        "\u00d7": " x ",
        "\u00f7": " / ",
        "\u221e": " infinity ",
    }
    for orig, rep in replacements.items():
        text = text.replace(orig, rep)
    return text


class ModelAnswersPDF(FPDF):
    def __init__(self, subject_title: str = "EXAMINATION"):
        super().__init__()
        import os
        if os.path.exists(r"C:\Windows\Fonts\nirmala.ttf"):
            self.add_font("nirmala", "", r"C:\Windows\Fonts\nirmala.ttf")
            self.add_font("nirmala", "B", r"C:\Windows\Fonts\nirmalab.ttf")
            self.add_font("nirmala", "I", r"C:\Windows\Fonts\nirmala.ttf")
        self.subject_title = sanitize_for_pdf(subject_title)

    def header(self):
        self.set_fill_color(15, 23, 42)
        self.rect(0, 0, 210, 14, "F")
        self.set_text_color(255, 255, 255)
        self.set_font("nirmala", "B", 9)
        self.set_xy(0, 2)
        self.cell(210, 10, "UNIVERSITY MODEL ANSWER SHEET -- EXTRACTA AI", align="C")
        self.ln(16)

    def footer(self):
        self.set_y(-14)
        self.set_font("nirmala", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")


def _strip_inline_md(text: str) -> str:
    """Remove inline markdown markers like **bold**, *italic*, and $LaTeX$ cleanly."""
    # Strip $...$ LaTeX (inline)
    text = re.sub(r'\$\$([^$]*)\$\$', lambda m: f'  {m.group(1).strip()}  ', text)
    text = re.sub(r'\$([^$\n]*)\$',   lambda m: f' {m.group(1).strip()} ',  text)
    # Strip bold/italic markers
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'\*([^*]+)\*',     r'\1', text)
    # Strip backtick code spans
    text = re.sub(r'`([^`]+)`', r'\1', text)
    return text


def _is_section_heading(line: str) -> bool:
    """Detect === SECTION === headings produced by the LLM."""
    return line.startswith("===") and line.endswith("===")


def write_markdown_to_pdf(pdf: FPDF, text: str):
    text = sanitize_for_pdf(text)
    lines = text.split("\n")
    in_table = False
    table_data = []
    line_h = 5

    pdf.set_font("nirmala", "", 10)
    pdf.set_text_color(30, 41, 59)

    def _flush_table():
        nonlocal in_table, table_data
        if table_data:
            _render_table(pdf, table_data)
        table_data = []
        in_table = False

    for line in lines:
        line = line.strip()

        if not line:
            if in_table:
                _flush_table()
            pdf.ln(2)
            continue

        # === SECTION HEADING ===
        if _is_section_heading(line):
            if in_table:
                _flush_table()
            heading = line.strip("= ").strip()
            pdf.set_font("nirmala", "B", 11)
            pdf.set_fill_color(230, 235, 248)
            pdf.set_text_color(40, 50, 120)
            pdf.set_x(pdf.l_margin)
            pdf.cell(0, 7, sanitize_for_pdf(heading), fill=True)
            pdf.ln(8)
            pdf.set_font("nirmala", "", 10)
            pdf.set_text_color(30, 41, 59)
            continue

        # Markdown ## / # headings
        if line.startswith("### "):
            if in_table: _flush_table()
            heading = _strip_inline_md(line[4:])
            pdf.set_font("nirmala", "B", 11)
            pdf.set_text_color(79, 70, 229)
            pdf.set_x(pdf.l_margin)
            pdf.cell(0, 7, sanitize_for_pdf(heading))
            pdf.ln(8)
            pdf.set_font("nirmala", "", 10)
            pdf.set_text_color(30, 41, 59)
            continue
        elif line.startswith("## ") or line.startswith("# "):
            if in_table: _flush_table()
            heading = _strip_inline_md(line[3:] if line.startswith("## ") else line[2:])
            pdf.set_font("nirmala", "B", 12)
            pdf.set_text_color(60, 80, 200)
            pdf.set_x(pdf.l_margin)
            pdf.cell(0, 8, sanitize_for_pdf(heading))
            pdf.ln(9)
            pdf.set_font("nirmala", "", 10)
            pdf.set_text_color(30, 41, 59)
            continue

        # Markdown Tables
        if line.startswith("|"):
            in_table = True
            parts = [p.strip() for p in line.split("|")[1:-1]]
            if parts and all(set(p) <= set("-: ") for p in parts):
                continue  # skip separator row
            table_data.append([_strip_inline_md(p) for p in parts])
            continue

        if in_table and not line.startswith("|"):
            _flush_table()

        # Numbered list: 1. 2. 3.
        if re.match(r'^\d+\.\s', line):
            clean = sanitize_for_pdf(_strip_inline_md(line))
            pdf.set_font("nirmala", "", 10)
            pdf.set_text_color(30, 41, 59)
            pdf.set_x(pdf.l_margin + 3)
            pdf.multi_cell(0, line_h, clean)
            pdf.ln(1)
            continue

        # Bullet points: - or *
        if re.match(r'^[\-\*\+]\s', line):
            clean = sanitize_for_pdf(_strip_inline_md(line[2:]))
            pdf.set_font("nirmala", "", 10)
            pdf.set_text_color(30, 41, 59)
            pdf.set_x(pdf.l_margin + 2)
            pdf.write(line_h, "-  ")
            pdf.multi_cell(0, line_h, clean)
            pdf.ln(1)
            continue

        # Bold-only lines (lines that were entirely **bold** headings from LLM)
        is_bold = re.match(r'^\*\*(.+)\*\*$', line)
        if is_bold:
            clean = sanitize_for_pdf(is_bold.group(1))
            pdf.set_font("nirmala", "B", 10)
            pdf.set_text_color(30, 41, 59)
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(0, line_h, clean)
            pdf.set_font("nirmala", "", 10)
            pdf.ln(1)
            continue

        # Regular paragraph text
        clean = sanitize_for_pdf(_strip_inline_md(line))
        pdf.set_font("nirmala", "", 10)
        pdf.set_text_color(30, 41, 59)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(0, line_h, clean)
        pdf.ln(1)

    if in_table and table_data:
        _flush_table()


def _render_table(pdf: FPDF, rows: List[List[str]]):
    """Render a markdown table with auto-height cells so text never overlaps."""
    if not rows:
        return

    col_count = max(len(r) for r in rows)
    page_w    = pdf.w - pdf.l_margin - pdf.r_margin
    col_w     = page_w / col_count
    line_h    = 5
    x_start   = pdf.l_margin

    for row_idx, row in enumerate(rows):
        padded = (row + [""] * col_count)[:col_count]
        sanitized = [sanitize_for_pdf(_strip_inline_md(str(c))) for c in padded]

        is_header = (row_idx == 0)
        if is_header:
            pdf.set_font("nirmala", "B", 9)
            pdf.set_fill_color(215, 225, 248)
            pdf.set_text_color(15, 23, 42)
        else:
            pdf.set_font("nirmala", "", 9)
            fill_r = (245, 247, 252) if row_idx % 2 == 0 else (255, 255, 255)
            pdf.set_fill_color(*fill_r)
            pdf.set_text_color(30, 41, 59)

        y_start = pdf.get_y()

        # Page-break guard before each row
        if y_start > pdf.h - pdf.b_margin - 25:
            pdf.add_page()
            y_start = pdf.get_y()

        # First pass: measure max lines needed in this row
        max_lines = 1
        for cell_text in sanitized:
            sw = pdf.get_string_width(cell_text)
            avail = col_w - 4
            n_lines = max(1, int(sw / max(avail, 1)) + 1)
            max_lines = max(max_lines, n_lines)
        row_h = max_lines * line_h + 3

        # Second pass: draw each cell at the same y_start, track max y
        max_y = y_start
        for ci, cell_text in enumerate(sanitized):
            x = x_start + ci * col_w
            pdf.set_xy(x, y_start)
            pdf.multi_cell(col_w, line_h, cell_text, border=1, align="L", fill=True)
            max_y = max(max_y, pdf.get_y())

        # Advance to the lowest cell bottom, add tiny gap
        pdf.set_xy(x_start, max_y)

    pdf.ln(4)


def compile_answers_pdf(results: List[Dict[str, Any]], header: Dict[str, str], task_id: str, pdf_path: str = "") -> str:
    subject = header.get("subject", "Academic Examination")
    pdf = ModelAnswersPDF(subject_title=subject)
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    # ── Title ──────────────────────────────────────────────────────
    pdf.set_font("nirmala", "B", 15)
    pdf.set_text_color(15, 23, 42)
    pdf.multi_cell(0, 10, sanitize_for_pdf(f"Model Answer Sheet: {subject}"), align="C")
    pdf.ln(3)

    # ── Metadata box (simple two-column rows using set_x + cell + ln) ──
    def meta_row(label: str, value: str):
        pdf.set_font("nirmala", "B", 9)
        pdf.set_text_color(71, 85, 105)
        pdf.set_x(18)
        pdf.cell(55, 6, sanitize_for_pdf(label))
        pdf.set_font("nirmala", "", 9)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(120, 6, sanitize_for_pdf(value))
        pdf.ln(6)

    y0 = pdf.get_y()
    pdf.set_fill_color(248, 250, 252)
    pdf.rect(15, y0, 180, 42, "F")

    meta_row("Subject:",          header.get("subject", "--"))
    meta_row("Paper Code:",       header.get("paper_code", "--"))
    meta_row("QP Code:",          header.get("qp_code", "--"))
    meta_row("Semester / Scheme:",
             f"{header.get('semester', '--')} / {header.get('scheme', '--')}")
    meta_row("Branch:",           header.get("branch", "--"))
    meta_row("Time & Marks:",
             f"{header.get('time', '--')}  |  Max Marks: {header.get('max_marks', '--')}")
    meta_row("Date:",
             header.get("date", "") or time.strftime("%d/%m/%Y"))

    pdf.ln(10)

    # ── Question Answer Blocks ──────────────────────────────────────
    for item in results:
        key    = item["key"]
        text   = item["text"]
        cat    = item["category"]
        ans    = item["answer"]
        q_type = item.get("q_type", "descriptive")
        options = item.get("options", [])

        # Question heading bar — amber for MCQ, default for others
        pdf.set_font("nirmala", "B", 11)
        if cat == "MCQ":
            pdf.set_fill_color(254, 243, 199)   # amber-50
            pdf.set_text_color(120, 60, 0)
        else:
            pdf.set_fill_color(241, 245, 249)
            pdf.set_text_color(15, 23, 42)
        heading = sanitize_for_pdf(f"Question {key}  [{cat.upper()}]")
        pdf.set_x(pdf.l_margin)
        pdf.cell(0, 8, heading, border="TB", align="L", fill=True)
        pdf.ln(9)

        # Question text in italics
        pdf.set_font("nirmala", "I", 10)
        pdf.set_text_color(71, 85, 105)
        pdf.multi_cell(0, 5, sanitize_for_pdf(f'"{text}"'))
        pdf.ln(3)

        # For MCQ: list the options before the answer
        if cat == "MCQ" and options:
            pdf.set_font("nirmala", "B", 9)
            pdf.set_text_color(120, 70, 10)
            pdf.cell(0, 6, "OPTIONS:")
            pdf.ln(6)
            for opt in options:
                pdf.set_font("nirmala", "", 9)
                pdf.set_text_color(30, 41, 59)
                pdf.set_x(pdf.l_margin + 6)
                pdf.cell(0, 5, sanitize_for_pdf(f"  ({opt['letter']})  {opt['text']}"))
                pdf.ln(5)
            pdf.ln(2)

        # Model Answer label
        pdf.set_font("nirmala", "B", 10)
        if cat == "MCQ":
            pdf.set_text_color(120, 70, 10)
            pdf.cell(0, 6, "ANSWER ANALYSIS:")
        else:
            pdf.set_text_color(79, 70, 229)
            pdf.cell(0, 6, "MODEL SOLUTION:")
        pdf.ln(7)

        # Answer content (markdown-aware)
        write_markdown_to_pdf(pdf, ans)
        pdf.ln(8)


    if pdf_path:
        base_name = os.path.basename(pdf_path)
        name_no_ext, _ = os.path.splitext(base_name)
        out_filename = f"model_answers_{name_no_ext}.pdf"
    else:
        out_filename = f"model_answers_{task_id}.pdf"

    out_path = os.path.join(OUTPUT_DIR, out_filename)
    pdf.output(out_path)
    return out_path

